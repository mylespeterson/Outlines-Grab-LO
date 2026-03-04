"""
Unit tests for course_outline_fetcher.py

Run with:
    python -m pytest test_course_outline_fetcher.py -v
"""

import io
import re
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch, call

import pytest

import course_outline_fetcher as cof


# ---------------------------------------------------------------------------
# extract_course_codes
# ---------------------------------------------------------------------------


class TestExtractCourseCodes:
    def test_finds_basic_codes(self):
        csv = "CET1023,CMP1117,NET2002\n"
        codes = cof.extract_course_codes(csv)
        assert "CET1023" in codes
        assert "CMP1117" in codes
        assert "NET2002" in codes

    def test_deduplicates(self):
        csv = "CET1023,CET1023,CET1023\n"
        codes = cof.extract_course_codes(csv)
        assert codes.count("CET1023") == 1

    def test_sorts_alphabetically(self):
        csv = "NET2002 CET1023 CMP1117"
        codes = cof.extract_course_codes(csv)
        assert codes == sorted(codes)

    def test_ignores_lowercase(self):
        csv = "cet1023,CET1023\n"
        codes = cof.extract_course_codes(csv)
        # lowercase version should not match; all returned codes must be uppercase
        assert all(c == c.upper() for c in codes)
        assert "CET1023" in codes

    def test_ignores_non_code_text(self):
        csv = "Hello, World! No codes here.\n"
        codes = cof.extract_course_codes(csv)
        assert codes == []

    def test_handles_empty_string(self):
        codes = cof.extract_course_codes("")
        assert codes == []

    def test_five_letter_prefix(self):
        csv = "MATH1234\n"
        codes = cof.extract_course_codes(csv)
        assert "MATH1234" in codes

    def test_two_letter_prefix(self):
        csv = "CS1234\n"
        codes = cof.extract_course_codes(csv)
        assert "CS1234" in codes

    def test_mixed_content(self):
        csv = (
            "Course,Code,Instructor\n"
            "Introduction to Networking,NET2002,Smith\n"
            "Computer Essentials,CMP1117,Jones\n"
            "Electronics Technology,CET1023,Williams\n"
        )
        codes = cof.extract_course_codes(csv)
        assert set(codes) == {"NET2002", "CMP1117", "CET1023"}


# ---------------------------------------------------------------------------
# group_by_prefix
# ---------------------------------------------------------------------------


class TestGroupByPrefix:
    def test_groups_correctly(self):
        codes = ["CET1023", "CET1024", "CMP1117", "NET2002"]
        groups = cof.group_by_prefix(codes)
        assert set(groups["CET"]) == {"CET1023", "CET1024"}
        assert set(groups["CMP"]) == {"CMP1117"}
        assert set(groups["NET"]) == {"NET2002"}

    def test_single_code(self):
        groups = cof.group_by_prefix(["CET1023"])
        assert groups == {"CET": ["CET1023"]}

    def test_empty_list(self):
        groups = cof.group_by_prefix([])
        assert groups == {}


# ---------------------------------------------------------------------------
# _parse_search_results
# ---------------------------------------------------------------------------


class TestParseSearchResults:
    def _make_html(self, entries):
        """Build a minimal HTML page with PDF links."""
        links = "\n".join(
            f'<tr><td>{code}</td><td><a href="{url}">{name}</a></td></tr>'
            for code, name, url in entries
        )
        return f"<html><body><table>{links}</table></body></html>"

    def test_finds_pdf_links(self):
        html = self._make_html([
            ("CET1023", "Electronics Technology", "/pdfs/CET1023.pdf"),
        ])
        results = cof._parse_search_results(html, "https://example.com")
        assert len(results) == 1
        assert results[0]["code"] == "CET1023"
        assert results[0]["pdf_url"] == "https://example.com/pdfs/CET1023.pdf"

    def test_absolute_pdf_url_unchanged(self):
        html = (
            '<html><body>'
            '<a href="https://other.example.com/CET1023.pdf">CET1023</a>'
            '</body></html>'
        )
        results = cof._parse_search_results(html, "https://example.com")
        assert results[0]["pdf_url"] == "https://other.example.com/CET1023.pdf"

    def test_ignores_non_pdf_links(self):
        html = (
            '<html><body>'
            '<a href="/page/CET1023.html">CET1023</a>'
            '</body></html>'
        )
        results = cof._parse_search_results(html, "https://example.com")
        assert results == []

    def test_deduplicates(self):
        html = (
            '<html><body>'
            '<a href="/pdfs/CET1023.pdf">CET1023</a>'
            '<a href="/pdfs/CET1023.pdf">CET1023</a>'
            '</body></html>'
        )
        results = cof._parse_search_results(html, "https://example.com")
        assert len(results) == 1

    def test_empty_html(self):
        results = cof._parse_search_results("<html></html>", "https://example.com")
        assert results == []

    def test_code_in_href(self):
        html = (
            '<html><body>'
            '<a href="/pdfs/CET1023.pdf">Download Outline</a>'
            '</body></html>'
        )
        results = cof._parse_search_results(html, "https://example.com")
        assert len(results) == 1
        assert results[0]["code"] == "CET1023"


# ---------------------------------------------------------------------------
# _clean_outcome_lines
# ---------------------------------------------------------------------------


class TestCleanOutcomeLines:
    def test_removes_page_numbers(self):
        lines = ["12", "Some real outcome here", "3"]
        cleaned = cof._clean_outcome_lines(lines)
        assert "12" not in cleaned
        assert "3" not in cleaned
        assert "Some real outcome here" in cleaned

    def test_removes_very_short_lines(self):
        lines = ["OK", "This is a valid outcome sentence"]
        cleaned = cof._clean_outcome_lines(lines)
        assert "OK" not in cleaned
        assert "This is a valid outcome sentence" in cleaned

    def test_empty_input(self):
        assert cof._clean_outcome_lines([]) == []

    def test_preserves_valid_lines(self):
        lines = [
            "Apply fundamental networking concepts to real-world scenarios",
            "Configure and troubleshoot TCP/IP networks",
        ]
        cleaned = cof._clean_outcome_lines(lines)
        assert cleaned == lines


# ---------------------------------------------------------------------------
# extract_learning_outcomes (with a synthetic PDF)
# ---------------------------------------------------------------------------


def _make_test_pdf(text: str) -> Path:
    """Create a real PDF with *text* using pdfplumber's underlying library."""
    try:
        from reportlab.pdfgen import canvas  # type: ignore
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        c = canvas.Canvas(tmp.name)
        y = 750
        for line in text.splitlines():
            c.drawString(50, y, line)
            y -= 15
            if y < 50:
                c.showPage()
                y = 750
        c.save()
        return Path(tmp.name)
    except ImportError:
        pytest.skip("reportlab not installed – skipping PDF generation test")


class TestExtractLearningOutcomes:
    def test_extracts_outcomes_after_header(self, tmp_path):
        pdf_path = _make_test_pdf(
            "Course Outline\n"
            "Learning Outcomes\n"
            "Apply networking concepts in practice\n"
            "Configure routers and switches\n"
            "Evaluation\n"
            "Midterm exam 40%\n"
        )
        outcomes = cof.extract_learning_outcomes(pdf_path)
        assert any("networking" in o.lower() for o in outcomes)
        # Should not include content past the stop header
        assert not any("midterm" in o.lower() for o in outcomes)

    def test_empty_if_no_lo_section(self, tmp_path):
        pdf_path = _make_test_pdf(
            "Course Outline\n"
            "Course Description\n"
            "This course covers networking.\n"
        )
        outcomes = cof.extract_learning_outcomes(pdf_path)
        assert outcomes == []

    def test_handles_missing_file(self, tmp_path):
        outcomes = cof.extract_learning_outcomes(tmp_path / "nonexistent.pdf")
        assert outcomes == []


# ---------------------------------------------------------------------------
# write_excel
# ---------------------------------------------------------------------------


class TestWriteExcel:
    def test_creates_file(self, tmp_path):
        results = [
            {
                "code": "CET1023",
                "name": "Electronics Technology",
                "outcomes": ["Apply circuit theory", "Use test equipment"],
                "error": "",
            }
        ]
        output = str(tmp_path / "test_output.xlsx")
        cof.write_excel(results, output)
        assert Path(output).exists()
        assert Path(output).stat().st_size > 0

    def test_excel_has_correct_sheets(self, tmp_path):
        from openpyxl import load_workbook

        results = [
            {"code": "CET1023", "name": "Test", "outcomes": ["Outcome 1"], "error": ""}
        ]
        output = str(tmp_path / "test_output.xlsx")
        cof.write_excel(results, output)

        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames
        assert "Learning Outcomes" in wb.sheetnames

    def test_summary_sheet_content(self, tmp_path):
        from openpyxl import load_workbook

        results = [
            {"code": "CET1023", "name": "Electronics", "outcomes": ["A", "B"], "error": ""},
            {"code": "CMP1117", "name": "Computers", "outcomes": [], "error": "Not found"},
        ]
        output = str(tmp_path / "test_output.xlsx")
        cof.write_excel(results, output)

        wb = load_workbook(output)
        ws = wb["Summary"]
        codes_in_sheet = [ws.cell(row=r, column=1).value for r in range(2, 4)]
        assert "CET1023" in codes_in_sheet
        assert "CMP1117" in codes_in_sheet

    def test_lo_sheet_lists_all_outcomes(self, tmp_path):
        from openpyxl import load_workbook

        outcomes = ["Apply circuit theory", "Use test equipment safely"]
        results = [
            {"code": "CET1023", "name": "Electronics", "outcomes": outcomes, "error": ""}
        ]
        output = str(tmp_path / "test_output.xlsx")
        cof.write_excel(results, output)

        wb = load_workbook(output)
        ws = wb["Learning Outcomes"]
        lo_values = [ws.cell(row=r, column=4).value for r in range(2, 4)]
        assert outcomes[0] in lo_values
        assert outcomes[1] in lo_values

    def test_handles_empty_results(self, tmp_path):
        output = str(tmp_path / "test_empty.xlsx")
        cof.write_excel([], output)
        assert Path(output).exists()

    def test_course_with_no_outcomes_shows_error(self, tmp_path):
        from openpyxl import load_workbook

        results = [
            {"code": "CET1023", "name": "Electronics", "outcomes": [], "error": "PDF download failed"}
        ]
        output = str(tmp_path / "test_output.xlsx")
        cof.write_excel(results, output)

        wb = load_workbook(output)
        ws = wb["Learning Outcomes"]
        row2_d = ws.cell(row=2, column=4).value
        assert "PDF download failed" in (row2_d or "")
