#!/usr/bin/env python3
"""
Course Outline Fetcher

Fetches course subjects from a Google Sheets spreadsheet, downloads
course outline PDFs from Cambrian College, extracts Learning Outcomes
and Objectives from each PDF, and writes everything to an Excel file.

Usage:
    python course_outline_fetcher.py

Requirements:
    pip install -r requirements.txt

The Google Sheets spreadsheet must be publicly accessible (shared with
"Anyone with the link can view").
"""

import io
import os
import re
import sys
import time
import tempfile
import logging
from pathlib import Path
from urllib.parse import urljoin

import requests
import pdfplumber
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEETS_ID = "1UAfay0aOwX3uHZSZ59SL5tFpUpnHcH5eqLZpJr_HsKo"
SHEETS_GID = "113402939"
SHEETS_CSV_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEETS_ID}"
    f"/export?format=csv&gid={SHEETS_GID}"
)

CAMBRIAN_BASE_URL = "https://cf.cambriancollege.ca"
CAMBRIAN_SEARCH_URL = (
    f"{CAMBRIAN_BASE_URL}/course_outlines/offerings/SearchPages/courseselect.cfm"
)
CAMBRIAN_TERM = "202526"

OUTPUT_FILE = "course_learning_outcomes.xlsx"

# Regex that matches course codes such as CET1023, CMP1117, NET2002, etc.
COURSE_CODE_RE = re.compile(r"\b([A-Z]{2,5})(\d{4})\b")

# Section headers that indicate the start of Learning Outcomes / Objectives
LO_HEADERS = re.compile(
    r"(learning\s+outcomes?|course\s+learning\s+outcomes?|"
    r"learning\s+objectives?|course\s+objectives?|outcomes?\s+and\s+objectives?)",
    re.IGNORECASE,
)

# Headers that signal the end of the LO section
LO_END_HEADERS = re.compile(
    r"(evaluation|assessment|grading|required\s+(text|material)|"
    r"schedule|weekly\s+schedule|course\s+schedule|outline|content|"
    r"resources|bibliography|methods?\s+of\s+instruction|delivery)",
    re.IGNORECASE,
)

REQUEST_DELAY = 1.0  # seconds between HTTP requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Step 1 – Fetch course codes from Google Sheets
# ---------------------------------------------------------------------------


def fetch_google_sheets_csv(url: str, timeout: int = 30) -> str:
    """Download the Google Sheets spreadsheet as CSV text.

    Args:
        url: The /export?format=csv URL for the spreadsheet.
        timeout: HTTP request timeout in seconds.

    Returns:
        The CSV text.

    Raises:
        requests.HTTPError: If the server returned a non-2xx status.
        requests.RequestException: For other network errors.
    """
    log.info("Fetching Google Sheets: %s", url)
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    return response.text


def extract_course_codes(csv_text: str) -> list[str]:
    """Return a sorted, deduplicated list of course codes found in *csv_text*.

    A course code is a sequence of 2-5 uppercase letters followed immediately
    by exactly 4 digits (e.g. CET1023, CMP1117, NET2002).

    Args:
        csv_text: Raw CSV content from the Google Sheets spreadsheet.

    Returns:
        Sorted list of unique course codes.
    """
    codes = set(COURSE_CODE_RE.findall(csv_text))
    # findall returns (prefix, number) tuples; join them back
    return sorted(f"{prefix}{number}" for prefix, number in codes)


def group_by_prefix(codes: list[str]) -> dict[str, list[str]]:
    """Group course codes by their alphabetic prefix.

    Args:
        codes: List of course codes.

    Returns:
        Dictionary mapping each prefix to the list of course codes that
        share that prefix.
    """
    groups: dict[str, list[str]] = {}
    for code in codes:
        match = COURSE_CODE_RE.match(code)
        if match:
            prefix = match.group(1)
            groups.setdefault(prefix, []).append(code)
    return groups


# ---------------------------------------------------------------------------
# Step 2 – Search Cambrian College and collect PDF URLs
# ---------------------------------------------------------------------------


def search_cambrian_for_prefix(
    session: requests.Session,
    prefix: str,
    term: str = CAMBRIAN_TERM,
) -> list[dict]:
    """Search the Cambrian course outline portal for all courses with *prefix*.

    Submits the search form with the given course prefix and term, then
    parses the results page to find course names and links to their
    outline PDFs.

    Args:
        session: A :class:`requests.Session` to reuse for all requests.
        prefix: The course code prefix to search for (e.g. ``"CET"``).
        term: The academic term code (e.g. ``"202526"``).

    Returns:
        A list of dicts, each containing:
        - ``"code"``  – course code (e.g. ``"CET1023"``)
        - ``"name"``  – course name as it appears on the page
        - ``"pdf_url"`` – absolute URL of the course outline PDF
    """
    log.info("Searching Cambrian for prefix=%s term=%s", prefix, term)

    payload = {
        "coursePrefix": prefix,
        "term": term,
        "year": "all",
    }

    time.sleep(REQUEST_DELAY)
    response = session.post(CAMBRIAN_SEARCH_URL, data=payload, timeout=30)
    response.raise_for_status()

    return _parse_search_results(response.text, response.url)


def _parse_search_results(html: str, base_url: str) -> list[dict]:
    """Parse the Cambrian search results HTML page.

    Looks for table rows or list items that contain a course code and a
    link to a PDF.

    Args:
        html: The full HTML text of the search results page.
        base_url: The URL of the page, used to resolve relative links.

    Returns:
        List of course info dicts (see :func:`search_cambrian_for_prefix`).
    """
    soup = BeautifulSoup(html, "lxml")
    results = []

    # Look for all anchor tags pointing to PDF files
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if not href.lower().endswith(".pdf"):
            continue

        pdf_url = urljoin(base_url, href)

        # Try to extract the course code from the link text or surrounding text
        text = anchor.get_text(strip=True)
        parent_text = anchor.parent.get_text(separator=" ", strip=True) if anchor.parent else text

        code_match = COURSE_CODE_RE.search(parent_text) or COURSE_CODE_RE.search(href)
        if not code_match:
            continue

        code = f"{code_match.group(1)}{code_match.group(2)}"
        name = text if text else code

        results.append({"code": code, "name": name, "pdf_url": pdf_url})

    # Deduplicate by (code, pdf_url)
    seen = set()
    unique = []
    for r in results:
        key = (r["code"], r["pdf_url"])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    return unique


# ---------------------------------------------------------------------------
# Step 3 – Download PDFs
# ---------------------------------------------------------------------------


def download_pdf(
    session: requests.Session,
    pdf_url: str,
    dest_dir: Path,
    filename: str,
) -> Path:
    """Download a PDF from *pdf_url* and save it to *dest_dir/filename*.

    Args:
        session: A :class:`requests.Session`.
        pdf_url: URL of the PDF to download.
        dest_dir: Directory in which to save the file.
        filename: The filename to use (should end with ``.pdf``).

    Returns:
        The :class:`~pathlib.Path` to the saved file.

    Raises:
        requests.HTTPError: If the download fails.
    """
    log.info("Downloading PDF: %s", pdf_url)
    time.sleep(REQUEST_DELAY)
    response = session.get(pdf_url, timeout=60, stream=True)
    response.raise_for_status()

    dest_path = dest_dir / filename
    with open(dest_path, "wb") as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)

    log.info("Saved %s (%d bytes)", dest_path, dest_path.stat().st_size)
    return dest_path


# ---------------------------------------------------------------------------
# Step 4 – Parse Learning Outcomes from PDF
# ---------------------------------------------------------------------------


def extract_learning_outcomes(pdf_path: Path) -> list[str]:
    """Extract Learning Outcomes and Objectives from a course outline PDF.

    Scans the PDF text for a section whose heading matches
    :data:`LO_HEADERS` and collects bullet points / numbered items until
    the next major section heading.

    Args:
        pdf_path: Path to the PDF file.

    Returns:
        A list of outcome/objective strings.  May be empty if none are
        found.
    """
    log.info("Parsing PDF: %s", pdf_path)
    outcomes: list[str] = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "\n".join(
                page.extract_text() or "" for page in pdf.pages
            )
    except Exception as exc:  # noqa: BLE001
        log.warning("Could not read PDF %s: %s", pdf_path, exc)
        return outcomes

    lines = full_text.splitlines()
    in_lo_section = False
    buffer: list[str] = []

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        if LO_HEADERS.search(stripped):
            in_lo_section = True
            buffer = []
            continue

        if in_lo_section:
            # Stop at the next major section heading
            if LO_END_HEADERS.search(stripped) and len(stripped) < 80:
                break
            buffer.append(stripped)

    if buffer:
        outcomes = _clean_outcome_lines(buffer)

    if not outcomes:
        log.warning("No learning outcomes found in %s", pdf_path)

    return outcomes


def _clean_outcome_lines(lines: list[str]) -> list[str]:
    """Filter and normalize raw lines from the LO section.

    Removes page headers/footers, very short lines, and lines that look
    like they are part of the document boilerplate.

    Args:
        lines: Raw lines extracted from the PDF.

    Returns:
        Cleaned list of outcome strings.
    """
    cleaned = []
    for line in lines:
        # Skip very short fragments
        if len(line) < 5:
            continue
        # Skip lines that look like page numbers
        if re.fullmatch(r"\d+", line):
            continue
        cleaned.append(line)
    return cleaned


# ---------------------------------------------------------------------------
# Step 5 – Write Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def write_excel(
    results: list[dict],
    output_path: str = OUTPUT_FILE,
) -> None:
    """Write all course learning outcomes to an Excel workbook.

    Creates two sheets:
    - **Summary**: one row per course with code, name, and number of
      outcomes found.
    - **Learning Outcomes**: one row per outcome, grouped by course.

    Args:
        results: List of dicts with keys ``"code"``, ``"name"``,
            ``"outcomes"`` (list of str), and optionally ``"error"``.
        output_path: Destination ``.xlsx`` file path.
    """
    wb = Workbook()

    # ---- Summary sheet ------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = ["Course Code", "Course Name", "# Learning Outcomes", "Notes"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_num, result in enumerate(results, start=2):
        ws_summary.cell(row=row_num, column=1, value=result["code"])
        ws_summary.cell(row=row_num, column=2, value=result.get("name", ""))
        ws_summary.cell(
            row=row_num, column=3, value=len(result.get("outcomes", []))
        )
        ws_summary.cell(
            row=row_num, column=4, value=result.get("error", "")
        )

    _autofit_columns(ws_summary)

    # ---- Learning Outcomes sheet --------------------------------------
    ws_lo = wb.create_sheet("Learning Outcomes")

    lo_headers = ["Course Code", "Course Name", "#", "Learning Outcome / Objective"]
    for col, header in enumerate(lo_headers, start=1):
        cell = ws_lo.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for result in results:
        code = result["code"]
        name = result.get("name", "")
        outcomes = result.get("outcomes", [])

        if not outcomes:
            # Still add a row so every course appears
            ws_lo.cell(row=row_num, column=1, value=code)
            ws_lo.cell(row=row_num, column=2, value=name)
            ws_lo.cell(row=row_num, column=3, value="")
            ws_lo.cell(
                row=row_num,
                column=4,
                value=result.get("error", "No learning outcomes found"),
            )
            row_num += 1
        else:
            for idx, outcome in enumerate(outcomes, start=1):
                ws_lo.cell(row=row_num, column=1, value=code if idx == 1 else "")
                ws_lo.cell(row=row_num, column=2, value=name if idx == 1 else "")
                ws_lo.cell(row=row_num, column=3, value=idx)
                ws_lo.cell(row=row_num, column=4, value=outcome)
                row_num += 1

    _autofit_columns(ws_lo)
    ws_lo.column_dimensions[get_column_letter(4)].width = 80

    wb.save(output_path)
    log.info("Wrote Excel output to %s", output_path)


def _autofit_columns(ws) -> None:
    """Set a reasonable column width for each column in *ws*."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def run(
    sheets_url: str = SHEETS_CSV_URL,
    output_path: str = OUTPUT_FILE,
    pdf_dir: Path | None = None,
) -> list[dict]:
    """Full pipeline: fetch → search → download → parse → write.

    Args:
        sheets_url: URL of the Google Sheets CSV export.
        output_path: Path of the output Excel file.
        pdf_dir: Directory to save downloaded PDFs.  If ``None``, a
            temporary directory is used and cleaned up afterwards.

    Returns:
        The list of result dicts that was written to the Excel file.
    """
    # ---- Fetch course codes -------------------------------------------
    try:
        csv_text = fetch_google_sheets_csv(sheets_url)
    except requests.RequestException as exc:
        log.error("Failed to fetch Google Sheets: %s", exc)
        sys.exit(1)

    codes = extract_course_codes(csv_text)
    if not codes:
        log.error("No course codes found in the spreadsheet. Exiting.")
        sys.exit(1)

    log.info("Found %d course code(s): %s", len(codes), ", ".join(codes))
    prefix_groups = group_by_prefix(codes)
    log.info("Grouped into %d prefix(es): %s", len(prefix_groups), ", ".join(prefix_groups))

    # ---- Set up PDF directory ----------------------------------------
    _tmp_dir = None
    if pdf_dir is None:
        _tmp_dir = tempfile.mkdtemp(prefix="cambrian_pdfs_")
        pdf_dir = Path(_tmp_dir)
        log.info("Saving PDFs to temporary directory: %s", pdf_dir)
    else:
        pdf_dir = Path(pdf_dir)
        pdf_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (compatible; CourseOutlineFetcher/1.0)"
    )

    # ---- Search, download, parse -------------------------------------
    # Build a map from course code -> course info dict
    course_map: dict[str, dict] = {code: {"code": code, "name": code, "outcomes": [], "error": ""} for code in codes}

    for prefix, prefix_codes in prefix_groups.items():
        try:
            search_results = search_cambrian_for_prefix(session, prefix)
        except requests.RequestException as exc:
            log.warning("Search failed for prefix %s: %s", prefix, exc)
            for code in prefix_codes:
                course_map[code]["error"] = f"Search failed: {exc}"
            continue

        log.info("Found %d course outlines for prefix %s", len(search_results), prefix)

        for item in search_results:
            code = item["code"]
            if code not in course_map:
                # Found in search but not originally in spreadsheet; skip
                continue

            course_map[code]["name"] = item["name"]

            safe_name = re.sub(r"[^A-Za-z0-9_-]", "_", code) + ".pdf"
            try:
                pdf_path = download_pdf(session, item["pdf_url"], pdf_dir, safe_name)
                outcomes = extract_learning_outcomes(pdf_path)
                course_map[code]["outcomes"] = outcomes
            except requests.RequestException as exc:
                log.warning("Failed to download PDF for %s: %s", code, exc)
                course_map[code]["error"] = f"PDF download failed: {exc}"
            except Exception as exc:  # noqa: BLE001
                log.warning("Failed to process %s: %s", code, exc)
                course_map[code]["error"] = f"Processing error: {exc}"

        # Mark any codes in this prefix group that had no search result
        found_codes = {r["code"] for r in search_results}
        for code in prefix_codes:
            if code not in found_codes and not course_map[code]["error"]:
                course_map[code]["error"] = "No course outline found on Cambrian website"

    # ---- Write Excel -------------------------------------------------
    results = sorted(course_map.values(), key=lambda r: r["code"])
    write_excel(results, output_path)

    # ---- Summary -----------------------------------------------------
    found = sum(1 for r in results if r["outcomes"])
    log.info(
        "Done. %d/%d courses have learning outcomes. Output: %s",
        found,
        len(results),
        output_path,
    )

    return results


def main() -> None:
    """Entry point when run from the command line."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Fetch Cambrian College course outlines and extract Learning Outcomes."
    )
    parser.add_argument(
        "--sheets-url",
        default=SHEETS_CSV_URL,
        help="Google Sheets CSV export URL (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT_FILE,
        help="Output Excel file path (default: %(default)s)",
    )
    parser.add_argument(
        "--pdf-dir",
        default=None,
        help="Directory to save downloaded PDFs (default: temporary directory)",
    )
    args = parser.parse_args()

    run(
        sheets_url=args.sheets_url,
        output_path=args.output,
        pdf_dir=Path(args.pdf_dir) if args.pdf_dir else None,
    )


if __name__ == "__main__":
    main()
